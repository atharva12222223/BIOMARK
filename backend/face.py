

import os, sys, cv2, numpy as np, sqlite3, time
from datetime import datetime

# ── Try importing openpyxl ──────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_OK = True
except ImportError:
    print("⚠  openpyxl not found. Run:  pip install openpyxl")
    print("   Excel saving will be skipped.\n")
    EXCEL_OK = False

# ── Try importing requests (for remote API mode) ────────
try:
    import requests
    REQUESTS_OK = True
except ImportError:
    REQUESTS_OK = False

# ══════════════════════════════════════════════════════
#  CONFIG  — adjust these paths if needed
# ══════════════════════════════════════════════════════
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH    = os.path.join(SCRIPT_DIR, "biomark.db")
FACE_DIR   = os.path.join(SCRIPT_DIR, "faces")
EXCEL_DIR  = os.path.join(SCRIPT_DIR, "attendance_sheets")
FACE_SIZE  = (100, 100)
THRESHOLD  = 0.20       # recognition confidence threshold
COOLDOWN   = 5          # seconds before same student can be marked again
CAM_INDEX  = 0          # camera index (try 1 if 0 doesn't work)

# ── REMOTE MODE ─────────────────────────────────────────
# Set these to sync attendance with a remote Render-hosted server.
# Leave REMOTE_API_URL empty ("") to use local SQLite only.
# Example: REMOTE_API_URL = "https://biomark.onrender.com"
REMOTE_API_URL = os.getenv("REMOTE_API_URL", "")
FACE_API_KEY   = os.getenv("FACE_API_KEY", "biomark-face-key-change-me")
USE_REMOTE     = bool(REMOTE_API_URL) and REQUESTS_OK

os.makedirs(EXCEL_DIR, exist_ok=True)

# ══════════════════════════════════════════════════════
#  DATABASE
# ══════════════════════════════════════════════════════
def get_db():
    if not os.path.exists(DB_PATH):
        print(f"❌ Database not found at: {DB_PATH}")
        print("   Make sure you run this from the backend/ folder.")
        sys.exit(1)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def get_student_by_regno(regno):
    db  = get_db()
    row = db.execute("SELECT id, name, regno, cls FROM students WHERE regno=?",
                     (regno.upper(),)).fetchone()
    db.close()
    return dict(row) if row else None

def mark_in_db(student_id, date):
    """Mark student as present. Returns True if newly marked, False if already present."""
    db = get_db()
    existing = db.execute(
        "SELECT present FROM attendance WHERE student_id=? AND date=?",
        (student_id, date)
    ).fetchone()

    already = existing and existing["present"] == 1

    if not already:
        db.execute(
            "INSERT INTO attendance (student_id, date, present) VALUES (?,?,1) "
            "ON CONFLICT(student_id, date) DO UPDATE SET present=1",
            (student_id, date)
        )
        db.commit()

    db.close()
    return not already  # True = freshly marked

def mark_remote(regno, date):
    """Mark attendance via remote API. Returns True if successful."""
    try:
        r = requests.post(
            f"{REMOTE_API_URL.rstrip('/')}/api/face/mark-attendance",
            json={"regno": regno.upper(), "api_key": FACE_API_KEY},
            timeout=5
        )
        return r.status_code == 200
    except Exception as e:
        print(f"  ⚠ Remote API error: {e}")
        return False

def get_all_students():
    db   = get_db()
    rows = db.execute("SELECT id, name, regno, cls FROM students ORDER BY name").fetchall()
    db.close()
    return [dict(r) for r in rows]

# ══════════════════════════════════════════════════════
#  EXCEL
# ══════════════════════════════════════════════════════
def get_excel_path(date_str):
    return os.path.join(EXCEL_DIR, f"attendance_{date_str}.xlsx")

def save_to_excel(name, regno, cls, date_str, time_str):
    if not EXCEL_OK:
        return

    path = get_excel_path(date_str)

    # Load existing or create new workbook
    if os.path.exists(path):
        wb = openpyxl.load_workbook(path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Attendance {date_str}"

        # ── Header row ──
        headers = ["#", "Name", "Register No", "Class", "Date", "Time", "Status"]
        header_fill = PatternFill("solid", fgColor="FFD700")
        header_font = Font(bold=True, color="1A253A", size=11)

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[1].height = 22

        # Column widths
        widths = [5, 24, 18, 16, 13, 10, 10]
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col)
            ].width = w

    # Check if already in file
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[2] == regno and row[4] == date_str:
            return  # already saved

    # New row
    next_row = ws.max_row + 1
    serial   = next_row - 1

    row_fill = PatternFill("solid", fgColor="F9F9F9" if serial % 2 == 0 else "FFFFFF")
    green_fill = PatternFill("solid", fgColor="D5F5E3")

    data = [serial, name, regno, cls, date_str, time_str, "Present"]
    for col, val in enumerate(data, 1):
        cell = ws.cell(row=next_row, column=col, value=val)
        cell.alignment = Alignment(horizontal="center" if col != 2 else "left",
                                   vertical="center")
        cell.fill = green_fill if col == 7 else row_fill
        if col == 7:
            cell.font = Font(bold=True, color="1E8449")

    # Thin border on all cells
    thin = Side(style="thin", color="DDDDDD")
    for col in range(1, 8):
        ws.cell(row=next_row, column=col).border = Border(
            left=thin, right=thin, top=thin, bottom=thin
        )

    ws.row_dimensions[next_row].height = 18
    wb.save(path)

# ══════════════════════════════════════════════════════
#  FACE RECOGNITION ENGINE
# ══════════════════════════════════════════════════════
cascade = cv2.CascadeClassifier(
    cv2.data.haarcascades + "haarcascade_frontalface_default.xml"
)

def preprocess(img):
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY) if len(img.shape) == 3 else img
    gray = cv2.resize(gray, FACE_SIZE)
    gray = cv2.equalizeHist(gray)
    return gray

def compare(face_a, face_b):
    a = preprocess(face_a)
    b = preprocess(face_b)

    ha = cv2.calcHist([a], [0], None, [256], [0, 256])
    hb = cv2.calcHist([b], [0], None, [256], [0, 256])
    cv2.normalize(ha, ha); cv2.normalize(hb, hb)
    hist  = max(0.0, cv2.compareHist(ha, hb, cv2.HISTCMP_CORREL))
    mse   = 1.0 / (1.0 + np.mean((a.astype(float) - b.astype(float))**2) / 500.0)
    diff  = 1.0 - np.mean(cv2.absdiff(a, b)) / 255.0
    return 0.4*hist + 0.35*mse + 0.25*diff

def load_faces():
    """Load all stored face images from disk. Returns {regno: [arrays]}"""
    stored = {}
    if not os.path.exists(FACE_DIR):
        return stored
    for regno in os.listdir(FACE_DIR):
        folder = os.path.join(FACE_DIR, regno)
        if not os.path.isdir(folder): continue
        imgs = []
        for f in sorted(os.listdir(folder)):
            if not f.endswith(".jpg"): continue
            img = cv2.imread(os.path.join(folder, f), cv2.IMREAD_GRAYSCALE)
            if img is not None: imgs.append(preprocess(img))
        if len(imgs) >= 5:
            stored[regno] = imgs
    return stored

def detect_and_crop(frame):
    """Returns (cropped_face, (x,y,w,h)) or (None, None)"""
    gray  = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    faces = cascade.detectMultiScale(gray, 1.1, 5, minSize=(60, 60))
    if len(faces) == 0: return None, None
    x, y, w, h = max(faces, key=lambda f: f[2]*f[3])
    return frame[y:y+h, x:x+w], (x, y, w, h)

def recognise(face_crop, stored):
    """Returns (regno, confidence) or (None, 0)"""
    best_regno = None
    best_score = 0.0
    for regno, imgs in stored.items():
        scores = sorted([compare(face_crop, s) for s in imgs], reverse=True)[:3]
        score  = sum(scores) / len(scores)
        if score > best_score:
            best_score = score
            best_regno = regno
    if best_score >= THRESHOLD:
        return best_regno, best_score
    return None, best_score

# ══════════════════════════════════════════════════════
#  DISPLAY HELPERS
# ══════════════════════════════════════════════════════
def draw_box(frame, x, y, w, h, color, label, sub_label=""):
    cv2.rectangle(frame, (x, y), (x+w, y+h), color, 2)

    # Background for text
    cv2.rectangle(frame, (x, y-46 if sub_label else y-26),
                  (x+w, y), color, -1)

    cv2.putText(frame, label, (x+6, y-28 if sub_label else y-8),
                cv2.FONT_HERSHEY_DUPLEX, 0.6, (0,0,0), 1, cv2.LINE_AA)
    if sub_label:
        cv2.putText(frame, sub_label, (x+6, y-8),
                    cv2.FONT_HERSHEY_DUPLEX, 0.5, (0,0,0), 1, cv2.LINE_AA)

def draw_status_bar(frame, text, color, alpha=0.7):
    h, w = frame.shape[:2]
    overlay = frame.copy()
    cv2.rectangle(overlay, (0, h-50), (w, h), color, -1)
    cv2.addWeighted(overlay, alpha, frame, 1-alpha, 0, frame)
    cv2.putText(frame, text, (20, h-18),
                cv2.FONT_HERSHEY_DUPLEX, 0.65, (255,255,255), 1, cv2.LINE_AA)

def draw_log(frame, log_entries):
    """Draw last 5 marked students in top-right corner."""
    h, w = frame.shape[:2]
    for i, entry in enumerate(log_entries[-5:]):
        y = 30 + i * 28
        cv2.putText(frame, f"✓ {entry}", (w-280, y),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 220, 0), 1, cv2.LINE_AA)

def draw_info(frame, stored_count, today_count, date_str):
    """Draw info overlay in top-left corner."""
    info_lines = [
        f"Date: {date_str}",
        f"Students loaded: {stored_count}",
        f"Marked today: {today_count}",
        "Q=quit  R=reload",
    ]
    for i, line in enumerate(info_lines):
        cv2.putText(frame, line, (10, 25 + i*22),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.5, (200, 200, 200), 1, cv2.LINE_AA)

# ══════════════════════════════════════════════════════
#  MAIN LOOP
# ══════════════════════════════════════════════════════
def main():
    print("\n" + "═"*54)
    print("  BIOMARK — Face Attendance System")
    print("═"*54)
    print(f"  DB   : {DB_PATH}")
    print(f"  Faces: {FACE_DIR}")
    print(f"  Excel: {EXCEL_DIR}")
    if USE_REMOTE:
        print(f"  🌐 Remote: {REMOTE_API_URL}")
    else:
        print("  🔒 Mode: Local only (set REMOTE_API_URL to sync)")
    print("  Keys : Q = quit  |  R = reload face data")
    print("═"*54 + "\n")

    # Load face data
    print("📂 Loading stored face images...")
    stored = load_faces()
    print(f"   Loaded {len(stored)} student(s): {', '.join(stored.keys()) or 'none'}\n")

    if not stored:
        print("⚠  No face data found. Register students first using port 8001.")
        print("   Continuing anyway — camera will open but no recognition will happen.\n")

    # Load student info lookup {regno: {name, cls, id}}
    students_info = {s["regno"]: s for s in get_all_students()}

    # Open camera
    print(f"📷 Opening camera (index {CAM_INDEX})...")
    cap = cv2.VideoCapture(CAM_INDEX)
    if not cap.isOpened():
        print(f"❌ Cannot open camera {CAM_INDEX}. Try changing CAM_INDEX at top of script.")
        sys.exit(1)

    cap.set(cv2.CAP_PROP_FRAME_WIDTH,  1280)
    cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 720)
    cap.set(cv2.CAP_PROP_FPS, 30)

    print("✅ Camera opened. Starting attendance marking...\n")

    # State
    cooldown_map  = {}     # {regno: last_marked_timestamp}
    marked_today  = set()  # regnos marked in this session
    log_entries   = []     # display log ["Name (regno)", ...]
    last_recog    = None   # last recognised result for stable display
    stable_frames = 0      # how many consecutive frames same person was seen
    STABLE_NEEDED = 8      # frames before marking (reduces false positives)

    today = datetime.now().strftime("%Y-%m-%d")

    while True:
        ret, frame = cap.read()
        if not ret:
            print("❌ Camera read failed."); break

        frame = cv2.flip(frame, 1)   # mirror for natural feel
        today = datetime.now().strftime("%Y-%m-%d")
        now   = time.time()

        # Detect face
        face_crop, bbox = detect_and_crop(frame)

        status_text  = "Scanning for face..."
        status_color = (60, 60, 60)
        box_color    = (180, 180, 180)
        box_label    = "Face detected"
        box_sub      = ""

        if face_crop is not None and bbox is not None:
            x, y, w, h = bbox

            # Recognise
            regno, confidence = recognise(face_crop, stored)

            if regno:
                student = students_info.get(regno.upper()) or students_info.get(regno)

                if student:
                    name = student["name"]
                    pct  = int(confidence * 100)

                    # Check cooldown
                    last_time = cooldown_map.get(regno, 0)
                    on_cooldown = (now - last_time) < COOLDOWN

                    # Stable recognition check
                    if last_recog == regno:
                        stable_frames += 1
                    else:
                        stable_frames = 0
                        last_recog    = regno

                    if on_cooldown:
                        # Already marked recently — show green confirmation
                        box_color    = (0, 200, 0)
                        box_label    = f"{name}"
                        box_sub      = f"Already marked ({pct}%)"
                        status_text  = f"✓ {name} — Already marked today"
                        status_color = (0, 130, 0)

                    elif stable_frames >= STABLE_NEEDED:
                        # MARK ATTENDANCE
                        # Mark in local DB
                        marked_new = mark_in_db(student["id"], today)
                        time_str   = datetime.now().strftime("%H:%M:%S")

                        # Also mark on remote server if configured
                        remote_ok = ""
                        if USE_REMOTE:
                            if mark_remote(regno, today):
                                remote_ok = " | Remote ✅"
                            else:
                                remote_ok = " | Remote ❌"

                        save_to_excel(
                            name, regno.upper(),
                            student.get("cls",""),
                            today, time_str
                        )

                        cooldown_map[regno] = now
                        marked_today.add(regno)
                        stable_frames = 0

                        log_entry = f"{name} ({regno.upper()})"
                        if log_entry not in log_entries:
                            log_entries.append(log_entry)

                        print(f"  ✅ {name:25s} | {regno.upper():12s} | {time_str} | conf:{pct}%"
                              + (" | DB updated" if marked_new else " | Already in DB") + remote_ok)

                        box_color    = (0, 255, 80)
                        box_label    = f"{name}"
                        box_sub      = f"MARKED ✓ ({pct}%)"
                        status_text  = f"✅ Attendance marked — {name}"
                        status_color = (0, 160, 0)

                    else:
                        # Recognised but stabilising
                        box_color    = (0, 200, 255)
                        box_label    = f"{name}"
                        box_sub      = f"Verifying... ({stable_frames}/{STABLE_NEEDED})"
                        status_text  = f"Verifying: {name} ({pct}% confidence)"
                        status_color = (180, 120, 0)

            else:
                # Face detected but not recognised
                last_recog    = None
                stable_frames = 0
                pct           = int(confidence * 100)
                box_color     = (0, 80, 255)
                box_label     = "Unknown"
                box_sub       = f"Conf: {pct}% (need {int(THRESHOLD*100)}%)"
                status_text   = f"Unknown face — confidence {pct}% (threshold {int(THRESHOLD*100)}%)"
                status_color  = (0, 60, 180)

            draw_box(frame, x, y, w, h, box_color, box_label, box_sub)

        else:
            last_recog    = None
            stable_frames = 0
            status_text   = "No face detected — please look at camera"
            status_color  = (50, 50, 50)

        # Draw overlays
        draw_status_bar(frame, status_text, status_color)
        draw_info(frame, len(stored), len(marked_today), today)
        draw_log(frame, log_entries)

        cv2.imshow("BIOMARK — Face Attendance  [Q=quit  R=reload]", frame)

        key = cv2.waitKey(1) & 0xFF
        if key == ord('q') or key == 27:  # Q or ESC
            break
        elif key == ord('r'):
            print("\n🔄 Reloading face data...")
            stored = load_faces()
            students_info = {s["regno"]: s for s in get_all_students()}
            print(f"   Loaded {len(stored)} student(s): {', '.join(stored.keys()) or 'none'}\n")

    cap.release()
    cv2.destroyAllWindows()

    # Final summary
    print("\n" + "═"*54)
    print(f"  Session ended — {len(marked_today)} student(s) marked")
    for r in sorted(marked_today):
        s = students_info.get(r.upper()) or students_info.get(r)
        if s: print(f"    ✓ {s['name']} ({r.upper()})")
    if EXCEL_OK and marked_today:
        print(f"\n  Excel saved to:")
        print(f"  {get_excel_path(today)}")
    print("═"*54 + "\n")

if __name__ == "__main__":
    main()